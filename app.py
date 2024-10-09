from flask import Flask, render_template, request, jsonify
import openpyxl
from openpyxl.styles import Font
import os

app = Flask(__name__)

# Path for static files including the image
app._static_folder = os.path.abspath("static")

excel_file = 'donation_list.xlsx'

if not os.path.exists(excel_file):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    headers = ['Donor Name', 'Food Name', 'Quantity', 'Date of Donation', 'Phone Number', 'Location       ', 'Address                 ']
    sheet.append(headers)
    
    for idx, cell in enumerate(sheet[1], start=1):  
        cell.font = Font(bold=True, color='0000FF')  
        sheet.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = len(cell.value) + 2 

    workbook.save(excel_file)

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/add_grocery', methods=['POST'])
def add_grocery():
    donor_name = request.form['donor_name']
    food_name = request.form['food_name']
    quantity = request.form['quantity']
    donate_date = request.form['donate_date']
    phone = request.form['phone']
    location = request.form['location']
    address = request.form['address']

    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active
    sheet.append([donor_name, food_name, quantity, donate_date, phone, location, address])
    workbook.save(excel_file)

    grocery_item = {
        'donor_name': donor_name,
        'food_name': food_name,
        'quantity': quantity,
        'donate_date': donate_date,
        'phone': phone,
        'location': location,
        'address': address
    }

    return jsonify(grocery_item)

@app.route('/get_groceries', methods=['GET'])
def get_groceries():
    workbook = openpyxl.load_workbook(excel_file)
    sheet = workbook.active

    donation_list = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        grocery_item = {
            'donor_name': row[0],
            'food_name': row[1],
            'quantity': row[2],
            'donate_date': row[3],
            'phone': row[4],
            'location': row[5],
            'address': row[6]
        }
        donation_list.append(grocery_item)

    return jsonify(donation_list)

if __name__ == '__main__':
    app.run(debug=True)
